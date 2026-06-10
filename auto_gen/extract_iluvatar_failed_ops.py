#!/usr/bin/env python3
"""Extract failed Iluvatar (Tianshu) operator names from the Excel results spreadsheet.

This script reads '第一批及格算子国产GPU测试.xlsx' and outputs the list of operators
that failed accuracy tests or were skipped, suitable for use with orchestrator.py --iluvatar.
"""

import argparse
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Error: 'openpyxl' is required. Install with: pip install openpyxl")
    sys.exit(1)


def extract_failed_ops(excel_path: str, output_path: str, sheet_name: str = None):
    """Extract failed operator names from Excel spreadsheet.

    An operator is considered "failed" if:
    - Its 精度结果 (accuracy result) column is NOT "精度通过"
    - OR its 状态 (status) column is "跳过"
    - OR the row is otherwise flagged as having issues

    Returns the list of unique operator names.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Try to find the main sheet
    if sheet_name:
        ws = wb[sheet_name]
    else:
        # Auto-detect: use the first sheet that contains operator data
        ws = None
        for sname in wb.sheetnames:
            sheet = wb[sname]
            for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str) and "算子" in cell:
                        ws = sheet
                        break
                if ws:
                    break
            if ws:
                break
        if ws is None:
            # Fallback to first sheet
            ws = wb[wb.sheetnames[0]]

    print(f"Using sheet: {ws.title}")

    # Find column indices from header row
    headers = [cell.value for cell in ws[1]]
    op_col = None
    result_col = None  # "天数测试结果" column

    for i, h in enumerate(headers):
        if h is None:
            continue
        h_str = str(h).replace("\n", "").strip()
        if "算子" in h_str and "名称" in h_str:
            op_col = i
        elif "天数测试结果" in h_str or "测试结果" in h_str:
            result_col = i
        elif "算子" in h_str or "op" in h_str.lower() or h_str.lower() == "operator":
            if op_col is None:
                op_col = i

    if op_col is None:
        print("Error: Could not find operator name column in spreadsheet header")
        print(f"Headers found: {headers}")
        sys.exit(1)

    if result_col is None:
        # Fallback: look for column that contains "失败"/"成功"/"跳过" values
        result_col = 2  # Default position based on observed spreadsheet layout

    print(f"Columns found: op_col={op_col}, result_col={result_col}")

    failed_ops = set()
    skipped_count = 0
    accuracy_fail_count = 0

    def is_numeric(s):
        """Check if a string represents a number (including floats)."""
        if not s:
            return False
        try:
            float(s)
            return True
        except ValueError:
            return False

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        op_name = row[op_col] if op_col < len(row) else None
        if not op_name or str(op_name).strip() == "" or str(op_name).strip() == "None":
            continue
        op_name = str(op_name).replace("\n", "").strip()

        # Strip aten:: prefix if present
        if op_name.startswith("aten::"):
            op_name = op_name[len("aten::"):]
        # Strip overload suffix (e.g. .Tensor, .Scalar)
        if "." in op_name:
            op_name = op_name.split(".")[0]

        result_val = str(row[result_col]).strip() if result_col is not None and result_col < len(row) else ""

        is_failed = False
        reason = ""

        if result_val == "失败":
            is_failed = True
            reason = "accuracy_fail"
            accuracy_fail_count += 1
        elif result_val == "跳过":
            is_failed = True
            reason = "skipped"
            skipped_count += 1
        elif result_val in ("成功", "", "None"):
            # Passed or empty/header row - skip
            pass
        elif is_numeric(result_val):
            # Numeric values (including benchmark speedup) indicate pass
            pass
        else:
            # Unknown status - treat as potential failure for safety
            is_failed = True
            reason = f"unknown_status({result_val})"
            accuracy_fail_count += 1

        if is_failed:
            failed_ops.add(op_name)
            print(f"  [{reason}] Row {row_idx}: {op_name}")

    wb.close()

    # Sort and write output
    sorted_ops = sorted(failed_ops)
    with open(output_path, "w") as f:
        f.write("\n".join(sorted_ops) + "\n")

    print(f"\nTotal failed operators extracted: {len(sorted_ops)}")
    print(f"  Accuracy failures: {accuracy_fail_count}")
    print(f"  Skipped: {skipped_count}")
    print(f"Output written to: {output_path}")

    return sorted_ops


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    default_excel = os.path.join(
        os.path.dirname(script_dir),
        "TianshuOperatorTest",
        "第一批及格算子国产GPU测试.xlsx",
    )
    default_output = os.path.join(script_dir, "ops_list_iluvatar.txt")

    parser = argparse.ArgumentParser(
        description="Extract failed Iluvatar (Tianshu) operators from Excel spreadsheet"
    )
    parser.add_argument(
        "-i", "--input",
        default=default_excel,
        help=f"Path to Excel spreadsheet (default: {default_excel})",
    )
    parser.add_argument(
        "-o", "--output",
        default=default_output,
        help=f"Output path for operator list (default: {default_output})",
    )
    parser.add_argument(
        "-s", "--sheet",
        default=None,
        help="Sheet name to use (default: auto-detect)",
    )
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"Error: Excel file not found: {args.input}")
        sys.exit(1)

    extract_failed_ops(args.input, args.output, args.sheet)


if __name__ == "__main__":
    main()