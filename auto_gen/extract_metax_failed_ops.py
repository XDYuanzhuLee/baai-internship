#!/usr/bin/env python3
"""
从 Excel 中提取沐曦（Metax）失败的算子列表。

Excel 文件格式：
- 列 C 的第 4 行起为"沐曦"结果
- 值："通过" 或 "失败"
- 列 A 为算子名称（aten::xxx 格式）
- 列 B 为 operator 类别

输出：去重后的纯算子名列表（去掉 aten:: 前缀和 overload 后缀）
"""

import argparse
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Error: 'openpyxl' is required. Install with: pip install openpyxl")
    sys.exit(1)


def extract_failed_ops(excel_path: str, output_path: str = None) -> list[str]:
    """Extract metax-failed operators from Excel."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    failed_ops = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Processing sheet: {sheet_name}")

        for row in ws.iter_rows(min_row=4, values_only=False):
            # Column A (1): operator name
            # Column B (2): category
            # Column C (3): metax result ("通过" or "失败")
            op_name = row[0].value if len(row) > 0 else None
            metax_result = row[2].value if len(row) > 2 else None

            if op_name is None:
                continue

            op_name = str(op_name).strip()
            if not op_name:
                continue

            # Check if metax result is "失败"
            if metax_result and str(metax_result).strip() == "失败":
                # Normalize: strip aten:: prefix
                if op_name.startswith("aten::"):
                    op_name = op_name[len("aten::"):]
                # Strip overload suffix
                if "." in op_name:
                    op_name = op_name.split(".")[0]
                if op_name and op_name not in failed_ops:
                    failed_ops.append(op_name)
                    print(f"  FAILED: {op_name}")

    failed_ops.sort()

    if output_path:
        with open(output_path, "w") as f:
            for op in failed_ops:
                f.write(f"{op}\n")
        print(f"\nSaved {len(failed_ops)} failed ops to {output_path}")

    print(f"\nTotal failed ops: {len(failed_ops)}")
    return failed_ops


def main():
    parser = argparse.ArgumentParser(description="Extract metax-failed operators from Excel")
    parser.add_argument(
        "excel_path",
        nargs="?",
        default=os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "operator_test_statistics",
            "第一批及格算子国产GPU测试.xlsx"
        ),
        help="Path to the Excel file"
    )
    parser.add_argument(
        "-o", "--output",
        default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "ops_list_metax.txt"),
        help="Output file path (default: ops_list_metax.txt)"
    )
    parser.add_argument("--print", action="store_true", help="Also print the list to stdout")

    args = parser.parse_args()

    if not os.path.exists(args.excel_path):
        print(f"Error: Excel file not found: {args.excel_path}")
        sys.exit(1)

    failed_ops = extract_failed_ops(args.excel_path, args.output)

    if args.print:
        print("\n--- Operator List ---")
        for op in failed_ops:
            print(op)


if __name__ == "__main__":
    main()
